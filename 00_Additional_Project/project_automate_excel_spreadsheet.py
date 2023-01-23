# In the transaction spreadsheet we have 3 columns transaction_id, product_id, price
# In this spreadsheet, we record all types of transactions, but because of some error
# the price here is wrong. Lets say we have to decrease the price by 10%.
# So if we have to do it manually, then we have to type a formula and make changes
# for all the rows in all the sheets. This manual process can be tedious.
# Lets automate this process (and also add a chart)

import openpyxl as xl # alias to make our code shorter
from openpyxl.chart import BarChart, Reference

def process_workbook(filename):

    wb = xl.load_workbook(filename)  #workbook object
    sheet = wb['Sheet1']
    cell = sheet['a1'] # access cells using coordinates
    cell = sheet.cell(1, 1) # access cells using cell method by passing row and column
    print(cell.value)

    # no of rows in the spreadsheet
    print(sheet.max_row)    # ans 4 rows

    ## iterate over the rows
    for row in range(2, sheet.max_row + 1):  # to include 4th row also, we add 1 and leave 1st header row
        cell = sheet.cell(row, 3)
        print(cell.value)
        corrected_price = cell.value * 0.9 #reduce the price by 10%
        # add corrected prices to the new column
        corrected_price_cell = sheet.cell(row, 4)
        corrected_price_cell.value = corrected_price    #update our spreadsheet

    ## add a chart (seelcting all values of 4th column)
    values = Reference(sheet,
                       min_row=2,
                       max_row=sheet.max_row,
                       min_col=4,
                       max_col=4)

    chart = BarChart()
    chart.add_data(values)
    sheet.add_chart(chart,'e2')

    newfile = 'transactions2.xlsx'
    wb.save(newfile)   #save all the changes on a new file


process_workbook('transactions.xlsx')