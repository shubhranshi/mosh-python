# The Command Query Separation Principle states that our methods or function should either be:
# commands that perform an action and change the state of a system or
# queries that return an answer to the caller and do not change the state of the system.
# So our methods should either be commands or queries but not both.

import openpyxl

wb = openpyxl.load_workbook("transaction2.xlsx")

wb.create_sheet("Sheet3")
# This is an example of a command method because it is responsible to perform a task. The task of creating a sheet.
# As result of calling this method the state of our system changes, in this case the workbook, changes.
# Every time we call this method it will create a new sheet.

# This is an example of a query method. We use it to access a given sheet.
sheet = wb["Sheet1"]
print(sheet)

# This is an example of a query method. We use it to access a given cell.
# However, this method violates the Command Query Separation Principle.
cell = sheet.cell(2, 3)
print(cell)

for row in range(1, 10):
    cell = sheet.cell(row, 1)
    print(cell.value)
# When run this for loop it will return the value of cell in column A and row 1 to 5.
# And for row 6 to 9 it will return "None"

# And after the for loop with the ".cell()" method the last row will be row 10 and not row 5.
# Because the ".cell()" method created the empty cell with the value "None"
# so if it does not find a cell, it will create empty cell --> This is command query separation violation
sheet.append([1, 2, 3])
# Therefore, this new row will get added at the 10th row. (This is command query separation violation)
# Ideally, if it is a query method, it should not change the state of the workbook.

wb.save("transaction3.xlsx")

# We can see the Command Query Separation Principle in action when trying to access an index that is not in the list.
numbers = ["a", "b", "c"]
# numbers[4]
# When we try to access the index 4 in this list. It will raise an exception,
# because that index does not exists. Instead of creating a new item.
# Here we are querying our list, so we should not change its state.
