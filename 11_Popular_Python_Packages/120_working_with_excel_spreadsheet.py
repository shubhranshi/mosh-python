# In this tutorial we are going to use the "openpyxl" package

import openpyxl

# When working with spreadsheets we need to start with a workbook object.
# We can either create an empty workbook in memory or load an existing workbook

# wb = openpyxl.Workbook() # Here we create a new workbook.

wb = openpyxl.load_workbook("transactions.xlsx") # Here we load an existing workbook
print(wb.sheetnames) # The "sheetnames" attribute returns a list with the name of each sheet in our workbook

sheet_1 = wb["Sheet1"] # To access a sheet in the workbook. And create a sheet object.

# Use the ".create_sheet()" method to create a new sheet in the workbook.
# The index parameter is the position of the sheet. "index 0" puts this sheet before sheet1
sheet_2 = wb.create_sheet(title="Sheet2", index=0)

# To create a cell object referencing a single cell
cell_a1 = sheet_1["a1"]

print(cell_a1.value) # Use the attribute ".value" to return the content in that cell. In this case it will return "transaction_id"
# can also change the value here.
cell_a1.value = 1
print(cell_a1.row) # Use the ".row" attribute to return the row number. In this case it will return "1"
print(cell_a1.column) # Use the ".column" attribute to retrun the column number. In this case it will return "1", not "A"
print(cell_a1.column_letter) # Use the ".column_letter" attribute to retrun the column letter. In this case it will return "A"
print(cell_a1.coordinate) # Use the coordinate attribute to retrun the coordinate of the cell, In this case it will return "A1".


# Another way to access a single cell is by using the ".cell()" method od the sheet object.
cell_b2 = sheet_1.cell(row=1, column=2)
print(cell_b2.value)

print(sheet_1.max_column)   # To get the number of columns.
print(sheet_1.max_row)  # To get the number of rows.

# To iterate over all the row and column of the sheet and get each value.
for row in range(1, sheet_1.max_row + 1):   # because the first row is 1 (not 0)
    for column in range(1, sheet_1.max_column + 1):
        cell = sheet_1.cell(row, column)
        print(cell.value)

column_a = sheet_1["a"]     # To access a columns in the sheet.
print(column_a)     # We will get a tuple with the cell objects in column A .

cell_a_c = sheet_1["a:c"]   # To access a range of columns
print(cell_a_c)     # Here we will get a tuple of tuples. On tuple for each column.

cell_2_4 = sheet_1["2"]     # To access a single rows
print(cell_2_4)     # Here we will get a tuple with the cells in row 2.

cell_2_4 = sheet_1["2:4"]   # same as ["a2:c4"], To access a range of rows
print(cell_2_4)     # Here we will get a tuple of tuples. On tuple for each row.

cell_a2_c4 = sheet_1["a2:c4"]   # To access a range of cells
print(cell_a2_c4)

sheet_1.append([1004, 4, 11.89]) # Use the append method to add a row to the end of the sheet.

# save the changes on a new spreadsheet.
wb.save("transaction2.xlsx")